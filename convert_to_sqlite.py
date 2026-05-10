"""
Conversion des fichiers XLSX vers une base SQLite pour AplusSurLeau.
Usage : python convert_to_sqlite.py
"""

import sqlite3
import openpyxl

DB_PATH = "aplusSurLeau.db"

FILES = {
    "xlsx/1_DB_Experte.xlsx": {
        "Disciplines":       "db_disciplines",
        "Progression":       "db_progression",
        "Matériel":          "db_materiel",
        "Spots":             "db_spots",
        "Règles de jointure": "db_regles_jointure",
    },
    "xlsx/1_Questionnaires.xlsx": {
        "M1 — Quelle discipline": "questionnaire_m1",
        "M2 — Cours et location": "questionnaire_m2",
        "Sheet3":                 "questionnaire_m3",
    },
    "xlsx/3-Regles_Association.xlsx": {
        "Règles d'association": "regles_association",
    },
}


def sanitize(name: str) -> str:
    """Transforme un nom de colonne en identifiant SQL valide."""
    return (
        name.strip()
        .lower()
        .replace(" ", "_")
        .replace("\n", "_")
        .replace("(", "")
        .replace(")", "")
        .replace("/", "_")
        .replace("'", "")
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("â", "a")
        .replace("ô", "o")
        .replace("î", "i")
        .replace("ù", "u")
        .replace("ç", "c")
        .replace("-", "_")
        .replace(".", "")
        .strip("_")
    )


def sheet_to_table(cursor, ws, table_name: str):
    headers_raw = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Ignorer les colonnes sans en-tête
    headers = [(i + 1, sanitize(h) if h else f"col_{i+1}") for i, h in enumerate(headers_raw)]

    # Dédoublonner les noms de colonnes
    seen = {}
    safe_headers = []
    for col_idx, h in headers:
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        safe_headers.append((col_idx, h))

    cols_def = ", ".join(f'"{h}" TEXT' for _, h in safe_headers)
    cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    cursor.execute(f'CREATE TABLE "{table_name}" ({cols_def})')

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, col_idx).value for col_idx, _ in safe_headers]
        if all(v is None for v in row):
            continue
        row = [str(v) if v is not None else None for v in row]
        placeholders = ", ".join("?" * len(row))
        cursor.execute(f'INSERT INTO "{table_name}" VALUES ({placeholders})', row)

    print(f"  ✓ Table '{table_name}' créée ({ws.max_row - 1} lignes)")


def main():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    for filename, sheets in FILES.items():
        print(f"\nTraitement de {filename}...")
        wb = openpyxl.load_workbook(filename)
        for sheet_name, table_name in sheets.items():
            if sheet_name not in wb.sheetnames:
                print(f"  ⚠ Onglet '{sheet_name}' non trouvé, ignoré.")
                continue
            ws = wb[sheet_name]
            sheet_to_table(cursor, ws, table_name)

    conn.commit()
    conn.close()
    print(f"\n✅ Base SQLite générée : {DB_PATH}")


if __name__ == "__main__":
    main()
